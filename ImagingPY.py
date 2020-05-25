# -*- coding: utf-8 -*-
"""
Created on Sun Apr  5 14:53:08 2020

@author: Theo.ROSSI
"""

import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import PySimpleGUI as sg
import configparser as cp
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os



class Tracker():
    
    def __init__(self):
        
        self.fig = fig
        self.ax1 = ax1
        self.ax2 = ax2
        self.ax3 = ax3
        self.ax4 = ax4
        self.ax5 = ax5
        
        self.cm = ['tomato', 'darkorange', 'turquoise', 'limegreen',
              'cornflowerblue', 'purple', 'magenta', 'darkred']
        
        self.xdata, self.ydata = [], []
        self.index = []
        self.index2 = []
        self.idx_fback = []
        self.ind = 0
        
        files = sorted(os.listdir(Path))
        
        LIST_ROI_FILE, LIST_COORD_FILE = [], []
        self.LIST_FRAME = []
        
        
        for file in files:
            
            if 'ROI.tif' in file:
                
                LIST_ROI_FILE.append(file)
                self.ROI_file = plt.imread('{}/{}'.format(Path, LIST_ROI_FILE[0]))
                self.ROI_file = self.ROI_file[:,:,0]
                self.im0 = ax1.imshow(self.ROI_file)
                
            elif '.coord' in file:
                
                LIST_COORD_FILE.append(file)
                Coord_file = pd.read_csv("{}/{}".format(Path, LIST_COORD_FILE[0]), sep="\t", header=1)
                
                X, Y = [],[]
                for i in range(len(Coord_file)): 
                    
                    basic_value = Coord_file.iloc[i,0]
                    
                    X.append(int(basic_value.split(',')[0]))
                    Y.append(int(basic_value.split(',')[1]))
                
                coordinates = {'X':X, 'Y':Y}
                self.df = pd.DataFrame(coordinates)
                
                
            elif '.ini' in file:
                
                ini = '{}/{}'.format(Path, file)
        
                config = cp.ConfigParser()
                config.read(ini)
                
                self.sampling_rate, self.px_dwell_time = float(config.get('_', 'lines.per.second')), float(config.get('_', 'pixel.dwell.time.in.sec'))*1000
            
            
            else:
                
                data = plt.imread('{}/{}'.format(Path, file)).transpose(1,0,2)
                data = data[:,1:-1,0]
                self.LIST_FRAME.append(data)
                
        self.time_vector = np.arange(0, 1.0/self.sampling_rate*len(data[1]), 1.0/self.sampling_rate)
        
        self.f0_start = np.ravel(np.where(self.time_vector >= 0.2))[0] 
        self.f0_stop = np.ravel(np.where(self.time_vector >= 0.45))[0] 
        
        self.frame = np.stack(i for i in self.LIST_FRAME).transpose(1,2,0)
        self.button = []
        
        self.slices = self.frame.shape[2]
        
        self.im = ax2.imshow(self.frame[:, :, self.ind])
        
        self.dF_F0 = []

        self.update()
        

    def markers(self, event):
        
        if event.inaxes in [ax1]:
            
            for idx in range(len(self.df)):
                
                if int(event.xdata) == self.df.loc[idx][0] and int(event.ydata) == self.df.loc[idx][1]:
                    
                    self.index.append(idx)
                    self.xdata.append(int(event.xdata))
                    self.ydata.append(int(event.ydata))
                    
                    self.ax1.scatter(self.xdata[-1], self.ydata[-1], s=30, marker='o')
                    self.ax2.scatter(15, self.index[-1], s=100, marker='>')
                    
                    self.ax1.annotate('Idx: {}'.format(idx),
                                      xy=(self.xdata[-1], self.ydata[-1]),
                                      xytext=(self.xdata[-1], self.ydata[-1]-5), color='w')
                    self.ax2.annotate('{}'.format(self.index[-1]), xy=(15,self.index[-1]), xytext=(15,self.index[-1]-8), color='w')
                    
                    if self.index != []:
                        
                        if '20Hz' in Path:
                            if '3stim' in Path:
                                self.stims = np.arange(0.5,0.64,0.05)
                            elif '10stim' in Path:
                                self.stims = np.arange(0.5,1,0.05)
                            
                        elif '50Hz' in Path:
                            if '3stim' in Path:
                                self.stims = np.arange(0.5,0.55,0.02)
                            if '10stim' in Path:
                                self.stims = np.arange(0.5,0.7,0.02)
                        
                        for stim in self.stims:
                            
                            self.ax4.scatter(stim, np.max(np.mean(self.frame[self.index[-1]-5 : self.index[-1]+5, :, self.ind], axis=0)), c='k', marker='v', linewidths=1)
                        
                        self.ax4.plot(self.time_vector, np.mean(self.frame[self.index[-1]-5 : self.index[-1]+5, :, self.ind], axis=0), alpha=0.5)
                    
                        
        if event.inaxes in [ax2]:
            
            if event.button == 2 and self.dF_F0 != []:
                
                self.index2.append(int(event.ydata))
                
                self.ax2.scatter(15, self.index2[-1], s=100, marker='>')
                
                self.button = np.mean(self.frame[self.index2[-1]-5 : self.index2[-1]+5, :, self.ind],
                                      axis=0)-np.mean(self.frame[self.idx_fback[-1]-3:self.idx_fback[-1]+3, :, self.ind], axis=0)
                
                self.f0 = np.mean(self.button[self.f0_start : self.f0_stop])
                
                self.dF_F0.append((self.button-self.f0)/self.f0)
                
                self.ax4.plot(self.time_vector, self.dF_F0[-1], alpha=0.5)
                
                if len(self.dF_F0) == self.slices:
                    
                    dataframe = pd.DataFrame(self.dF_F0, index=None, columns=None).transpose()
                    dataframe['Average'] = np.mean(self.dF_F0, axis=0)
                    dataframe['Time'] = self.time_vector
                    DATAFRAME.append(dataframe)

                    sg.popup_ok('Dataframe ready to be stored')
                
            if event.button == 3:
                
                self.idx_fback.append(int(event.ydata))
                
                for value in range(len(self.idx_fback)):
                    
                    self.ax2.scatter(15, int(event.ydata), s=100, c=self.cm[value], marker='$f$')
                    
                    self.ax3.plot(self.time_vector, np.mean(self.frame[self.idx_fback[-1]-3:self.idx_fback[-1]+3, :, self.ind], axis=0),
                                  c=self.cm[value], alpha=0.4)
    
    
    def presskey(self, event):
        
        if event.key == 'd':
            
            self.idx_fback.pop()
        
        elif event.key == 'D':
            
            if len(self.index) == 1:
                self.index2.pop()
            
            else:
                self.index.pop()
                
            self.xdata.pop()
            self.ydata.pop()
        
        elif event.key == 'c':
            
            self.ax1.clear()
            self.ax2.clear()
            self.ax3.clear()
            self.ax4.clear()
            self.ax5.clear()
            
            self.im0 = ax1.imshow(self.ROI_file)
            
            self.xdata, self.ydata = [], []
            self.index = []
            self.index2 = []
            self.idx_fback = []
            self.button = []
            
            self.im = ax2.imshow(self.frame[:, :, self.ind])
            
            self.dF_F0 = []
            
        
        elif event.key == 'enter':
            
            self.idx_fback = [self.idx_fback[-1]]
            self.ax2.clear()
            self.ax3.clear()

            self.button = np.mean(self.frame[self.index[-1]-5 : self.index[-1]+5, :, self.ind],
                                  axis=0)-np.mean(self.frame[self.idx_fback[-1]-3:self.idx_fback[-1]+3, :, self.ind], axis=0)
            
            self.ax3.plot(self.time_vector, np.mean(self.frame[self.idx_fback[-1]-3:self.idx_fback[-1]+3, :, self.ind], axis=0),
                              c=self.cm[0], alpha=0.4, label='Raw')
        
        elif event.key == 'm':
            
            self.index2.pop()
            self.dF_F0.pop()
        
        elif event.key == '+':
            
            self.ax4.clear()
            
            self.button = np.mean(self.frame[self.index[-1]-5 : self.index[-1]+5, :, self.ind],
                                  axis=0)-np.mean(self.frame[self.idx_fback[-1]-3:self.idx_fback[-1]+3, :, self.ind], axis=0)
            
            self.f0 = np.mean(self.button[self.f0_start : self.f0_stop])
            
            self.dF_F0.append((self.button-self.f0)/self.f0)
            
            self.ax4.plot(self.time_vector, self.dF_F0[-1])
            
            self.ax4.set_ylabel('dF/F0')
            self.ax5.legend('{}'.format(len(self.dF_F0)))
        
                
    def onscroll(self, event):
        
        if event.button == 'up':
            
            self.ind = (self.ind + 1) % self.slices
            
        else:
            
            self.ind = (self.ind - 1) % self.slices

        self.update()

        
    def update(self):
        
        self.ax1.clear()
        self.ax2.clear()
        self.ax3.clear()
        self.ax4.clear()
        self.ax5.clear()
        
        self.im0 = ax1.imshow(self.ROI_file)
        self.im = ax2.imshow(self.frame[:, :, self.ind])
        
        for x, y, idx in zip(range(len(self.xdata)), range(len(self.ydata)), range(len(self.index))):
            
            self.ax1.scatter(self.xdata[x], self.ydata[y], s=30, marker='o')
            self.ax1.annotate('Idx: {}'.format(self.index[idx]),
                                      xy=(self.xdata[x], self.ydata[y]),
                                      xytext=(self.xdata[x], self.ydata[y]-5), color='w')
           
            self.ax2.scatter(15, self.index[idx], s=100, marker='>')
            self.ax2.annotate('{}'.format(self.index[idx]),
                              xy=(15,self.index[idx]),
                              xytext=(15,self.index[idx]-8), color='w')
            
        for idx2 in range(len(self.index2)):
            
            self.ax2.scatter(15, self.index2[idx2], s=100, marker='>')
        
        for value in range(len(self.idx_fback)):
            
            self.ax2.scatter(15, self.idx_fback[value], s=100, c=self.cm[value], marker='$f$')
            self.ax3.plot(self.time_vector, np.mean(self.frame[self.idx_fback[value]-3:self.idx_fback[value]+3, :, self.ind], axis=0),
                          c=self.cm[value], alpha=0.4)
        
        if self.dF_F0 == [] and self.index != []:
                
            self.ax4.plot(self.time_vector, np.mean(self.frame[self.index[-1]-5 : self.index[-1]+5, :, self.ind], axis=0), alpha=0.5, label='Raw')
            
            for stim in self.stims:
                
                self.ax4.scatter(stim, np.max(np.mean(self.frame[self.index[-1]-5 : self.index[-1]+5, :, self.ind], axis=0)), c='k', marker='v', linewidths=1)
        
            if self.button != []:
                
                self.ax4.plot(self.time_vector, np.mean(self.frame[self.index[-1]-5 : self.index[-1]+5, :, self.ind],
                            axis=0)-np.mean(self.frame[self.idx_fback[-1]-3:self.idx_fback[-1]+3, :, self.ind],
                            axis=0), 'b', alpha=0.5, label='Raw-fback')
            
            self.ax4.legend()
            
        if self.dF_F0 != []:
            
            self.ax4.plot(self.time_vector, self.dF_F0[-1])
                
            for profile in range(len(self.dF_F0)):
                    
                self.ax5.plot(self.time_vector, self.dF_F0[profile], alpha=0.3)
                self.ax5.plot(self.time_vector, np.mean(self.dF_F0, axis=0), 'r')
            
            for stim in self.stims:
                
                self.ax4.scatter(stim, np.max(self.dF_F0[-1]), c='k', marker='v', linewidths=1)
                self.ax5.scatter(stim, np.max(self.dF_F0), c='k', marker='v', linewidths=1)
    
    
        self.ax1.set_title('{}\n ROI frame'.format(Path), fontsize=8)
        self.ax2.set_title('Linescan #{}'.format(self.ind), fontsize=20)
        self.ax3.set_title('Background noise')
        self.ax3.set_ylim(0,50)
        self.ax3.grid(axis='y', linestyle='--')
        
        
        if self.index != []:
            
            for i in range(len(self.index)):
                
                self.ax4.set_title('STP button {}'.format(self.index[-1]))
                
        else:
            
             self.ax4.set_title('STP button {}'.format(self.index))
             
        self.ax4.set_ylabel('dF/F0')
        self.ax4.grid(axis='y', linestyle='--')
        
        self.ax5.set_title('STP dF_F0')
        self.ax5.set_xlabel('Time (s)')
        self.ax5.grid(axis='y', linestyle='--')
        self.ax5.legend('{}'.format(len(self.dF_F0)))
        
        self.im.axes.figure.canvas.draw()
       
                
if __name__ == '__main__':
     
    Path = []
    DATAFRAME = []
    file_excel = []
    
    sg.theme('DarkBlue')
    
    layout = [[sg.Text('Folder Name'), sg.InputText(), sg.FolderBrowse(), sg.Button('Start Folder')],
              [sg.Text('Excel Name'), sg.InputText(size=(40,1)), sg.Text('Sheet'), sg.InputText(size=(7,1)), sg.Button('To Excel')],
              [sg.Button('Clear', button_color=('white','darkred'))]]
    
    window = sg.Window('Load files', layout, location=(0,0))
    
    while True:
        event, value = window.read()
        
        try:
            if event in (None, 'Close'):
                
                plt.close()
                file_excel.clear()
                DATAFRAME.clear()
                break
            
            if event == 'Start Folder':
                
                Path = value[0]
                
                fig = plt.figure(figsize=(19,10), tight_layout=True)
                gs = gridspec.GridSpec(3,3)
                ax1 = fig.add_subplot(gs[:,0])
                ax2 = fig.add_subplot(gs[:,1])
                ax3 = fig.add_subplot(gs[0,2])
                ax4 = fig.add_subplot(gs[1,2])
                ax5 = fig.add_subplot(gs[2,2])
                
                tracker = Tracker()
                fig.canvas.mpl_connect('scroll_event', tracker.onscroll)
                fig.canvas.mpl_connect('button_release_event', tracker.markers)
                fig.canvas.mpl_connect('key_release_event', tracker.presskey)
            
            
            if event == 'To Excel':
                
                filepath = 'E:'+str(value[1])+'.xlsx'
                
                if file_excel == []:
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = str(value[2])
                    
                    for r in dataframe_to_rows(DATAFRAME[0], index=False, header=True):
                        ws.append(r)
                        
                    wb.save(filepath)
                    
                    file_excel.append(filepath)
                    
                else:
                        
                    wb = load_workbook(filepath)
                    sheet = wb.create_sheet(str(value[2]))
                    
                    for r in dataframe_to_rows(DATAFRAME[0], index=False, header=True):
                        sheet.append(r)
                    
                    wb.save(filepath)
                
                sg.popup_ok('Profiles have been stored in: {}'.format(value[1]))
                
                DATAFRAME.clear()
            
            if event == 'Clear':
                
                Path = []
                DATAFRAME.clear()
                file_excel.clear()
                plt.close()
                
        except:
            pass
    
    window.close()